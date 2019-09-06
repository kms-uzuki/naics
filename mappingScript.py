# Script created in order to streamline quality control checking of CEUS data.
# Only works with specific files, but you can probably understand what's going on here.
# Far less optimized than it could be

import openpyxl
import time
from collections import defaultdict

start1 = time.time()

# assigns a building type (sector) to each site based on either our code or the surveyor's code.
def cat_assign(beg, end, sheet_1, sheet_2, _book, _file):
    
    map_dict = defaultdict(str)


    for i in range(93, 1038): # specific to the file we have internally, please change if that file changes
        if str(sheet_2.cell(row = i, column = 6).value) == 'Commercial':
            b = sheet_2.cell(row = i, column = 2).value
            h = str(sheet_2.cell(row = i, column = 8).value)
            map_dict[b] = h


    keyy = {
        'OFFICE': "Office",
        'RESTAURANT': "Restaurant",
        'FOOD/LIQUOR': "Food Stores",
        'RETAIL STORE': "Retail",
        'WAREHOUSE': "Warehouse",
        'HEALTH CARE': "Health Care",
        'HOTEL': "Lodging",
        'REFR WAREHOUSE': "Refrigerated Warehouse",
        'COLLEGE': "College",
        'SCHOOL': "School",
        'MISC': "Miscellaneous"
        }
    for i in range(beg, end + 1): # if col. 31 is empty, we pull the NAICS code and building type from the surveyor
        if sheet_1.cell(i, 31).value == None:
            temp = map_dict[str(sheet_1.cell(i, 27).value)]
            sheet_1.cell(i, 33).value = keyy.get(temp)
        else:
            if str(sheet_1.cell(i, 31).value).upper() == 'INC': # if inc, we pull the NAICS code and building type from the utility
                t = map_dict[str(sheet_1.cell(i, 24).value)]
                sheet_1.cell(i, 33).value = keyy.get(t)
            else:
                temp = map_dict[str(sheet_1.cell(i, 32).value)] # else, pull building type from verified/new code
                sheet_1.cell(i, 33).value = keyy.get(temp)
    
    _book.save(_file)

# creates the color scheme that allows for easier visual comparison to ADM's building types.
def colorer(begin, end, _sheet, _book, _file): 
    
    for i in range(begin, end + 1):

        _cell = "AG{}".format(i)
        # non-commercial codes
        if str(_sheet.cell(i, 31).value).lower() != "inc" and _sheet.cell(i, 33).value == None:
            _sheet[_cell].fill = openpyxl.styles.GradientFill(stop = ("00FFE4", "FF00E5"))

        # inc, no utility code
        elif str(_sheet.cell(i, 31).value).lower() == "inc" and _sheet.cell(i, 33).value == None:
            _sheet[_cell].fill = openpyxl.styles.PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")

        # 1 codes, not done codes; includes inc with utility codes
        elif _sheet.cell(i, 31).value == None or str(_sheet.cell(i, 31).value).lower() == "inc":
            if _sheet.cell(i, 33).value in _sheet.cell(i, 10).value: # our building type matches ADM's
                if "Office" in _sheet.cell(i, 33).value: # office buildings get special color for match
                    _sheet[_cell].fill = openpyxl.styles.PatternFill(start_color = "F79646", end_color = "F79646", fill_type = "solid")
                else: # match between surveyor's code and ADM's code
                    _sheet[_cell].fill = openpyxl.styles.PatternFill(start_color = "76933C", end_color = "76933C", fill_type = "solid")
            else: #our building type doesn't match ADM's
                _sheet[_cell].fill = openpyxl.styles.PatternFill(start_color = "C4BD97", end_color = "C4BD97", fill_type = "solid")

        # all done codes
        elif _sheet.cell(i, 31).value != None:
            if "Office" == str(_sheet.cell(i, 33).value) and str(_sheet.cell(i, 33).value) in str(_sheet.cell(i, 10).value): # offices match
                _sheet[_cell].fill = openpyxl.styles.PatternFill(start_color = "00FAFF", end_color = "00FAFF", fill_type = "solid")
            else:
                if _sheet.cell(i, 33).value != None and str(_sheet.cell(i, 33).value) not in str(_sheet.cell(i, 10).value): # building type mismatch
                    _sheet[_cell].fill = openpyxl.styles.PatternFill(start_color = "FFE500", end_color = "FFE500", fill_type = "solid")
                elif str(_sheet.cell(i, 33).value) == str(_sheet.cell(i, 10).value): # match between our code and ADM's code
                    _sheet[_cell].fill = openpyxl.styles.PatternFill(start_color = "00FF00", end_color = "00FF00", fill_type = "solid")
    _book.save(_file)


if __name__ == "__main__":
    bookName = input("File to edit (full path): ")
    bookSheet = input("Excel worksheet name with data: ")
    mapFile = input("Mapping table (full path, specific file should be NAICS to SECTOR 2018.xlsx): ")
    mapSheet = input("Mapping sheet with data (should be Electricity 2017): ")
    _begin = int(input("Row start: "))
    _end = int(input("Row end: "))
    
    start2 = time.time()
    zzz = openpyxl.load_workbook(bookName)
    _sheet1 = zzz[bookSheet]
    yyy = openpyxl.load_workbook(mapFile)
    _sheet2 = yyy[mapSheet]
    
    start3 = time.time()
    cat_assign(_begin, _end, _sheet1, _sheet2, zzz, bookName)
    start4 = time.time()
    colorer(_begin, _end, _sheet1, zzz, bookName)

    end = time.time()
    
    elapsed1 = end - start1
    elapsed2 = end - start2
    elapsed3 = start3 - start2
    elapsed4 = end - start4
    print("\nFinished\nTotal elapsed: {}\nMain elapsed: {}\nFn 1 elapsed: {}\nFn 2 elapsed: {}".format(elapsed1, elapsed2, elapsed3, elapsed4))
