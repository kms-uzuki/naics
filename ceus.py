# Script created in order to streamline quality control checking of CEUS data.
# Data collected by ADM Energy, analyzed by the CEC
# TODO: separate chain of events for when NAICS_site_selected is in set { 531120, 531190, 531312 }.
    ## Need to analyze these sites separately and potentially assign two NAICS codes.

######################################################################################################
###### NOTE: If you prefer Chrome not autoclosing every time you quit or the program autosaves, ######
########### comment out <import os> as well as <os.system('taskkill /im chrome.exe /f')>. ############
######################################################################################################

import openpyxl    # main library, allows for diverse read/write functions on Excel files
import webbrowser  # self-explanatory
import urllib      # url formatting
import time        # used to make pages open in a certain order
import collections # defaultdict container is somewhat more useful than the standard dictionary
import os          # allows for chrome window force-close (I'm lazy)
import threading   # used for optimization testing
    
def naics_web_lookup(naics, address):
    
    # google search queries
    a = ['https://www.google.com/search?q={}', 
    'https://www.google.com/maps/search/?api=1&query={}'] 
    for item in a:
        addr_url = urllib.parse.urlparse(item.format(address))
        webbrowser.open_new_tab(addr_url.geturl())
        time.sleep(0.01)
        
    # census.gov naics queries
    for j in naics:
        if j != 'NA' and j != None:
            naics_url = urllib.parse.urlparse('https://www.census.gov/cgi-bin/sssd/naics/naicsrch?input={}&search=2017+NAICS+Search&search=2017'.format(str(j)))
            webbrowser.open_new_tab(naics_url.geturl())
        time.sleep(0.01)

# helper function if entering a new code
def checker(code, code_list):
    for codes in code_list:
        if code == codes:
            return code
    return 0

# dictionary of valid NAICS codes, requires a document you can find on the NAICS site 
# https://www.census.gov/eos/www/naics/2017NAICS/2017_NAICS_Structure.xlsx
def codeListGenerator(_list, _bookname):
    for j in range(4, 2219):
        x = _bookname.cell( j,  2).value
        _list.append(x)
    return _list

# dictionary of all lower-level NAICS code descriptions, requires a document you can find on the NAICS site
# https://www.census.gov/eos/www/naics/2017NAICS/2017_NAICS_Index_File.xlsx
def naicsDictionary(_dict, _bookname):
    for k in range(2, 20059):
        a = _bookname.cell( k,  1).value
        b = str(_bookname.cell( k,  2).value)
        _dict[a].append(b)
        sorted(_dict.keys())
    return _dict

# main function to read/write
def ceushelper(begin, end, editorFile, structureFile, indexFile):
    
    #load your file to edit
    book = openpyxl.load_workbook(editorFile, data_only = True)
    sheet = book.active
    
    #load your file of NAICS codes
    book_2 = openpyxl.load_workbook(structureFile)
    sheet_2 = book_2['Sheet1']
    
    #load your file of NAICS code descriptions
    book_3 = openpyxl.load_workbook(indexFile)
    sheet_3 = book_3.active

    code_list = [] # for NAICS codes
    naics_desc = collections.defaultdict(list) # for code descriptions
    
    st = time.time()
    # codeListGenerator(code_list, sheet_2)
    # naicsDictionary(naics_desc, sheet_3)
    
    ## threading here saves maybe half a second max
    ##but we take those
    aa = threading.Thread(target = codeListGenerator, args = (code_list, sheet_2))
    bb = threading.Thread(target = naicsDictionary, args = (naics_desc, sheet_3))
    aa.start()
    bb.start()
    aa.join()
    bb.join()
    print("\n{}\n".format(time.time() - st))

    save_counter = 0 # used for autosave
    
    for i in range(begin, end):
        save_action = 0 # used for manual save
        
        # NAICS codes assigned by utility, ADM surveyor, etc. Change as necessary.
        naics_util = sheet.cell(i, 24).value 
        naics_smpl = sheet.cell(i, 25).value 
        naics_svyr = sheet.cell(i, 26).value 
        naics_sele = sheet.cell(i, 27).value 
        naics_call = sheet.cell(i, 28).value 
        naics_supp = sheet.cell(i, 29).value
        # use set here to ensure only distinct codes, python sets won't necessarily preserve order
        naics_set = { naics_util, naics_smpl, naics_svyr, naics_sele, naics_call, naics_supp }

        # matches input number to code
        da = {
            "1": naics_util,
            "2": naics_smpl,
            "3": naics_svyr,
            "4": naics_sele,
            "5": naics_call,
            "6": naics_supp
            }
        
        # terribly implemented because I'm lazy, matches input to a command
        dd = {
            "1": "VER",
            "2": "VER",
            "3": "VER",
            "4": "VER",
            "5": "VER",
            "6": "VER",
            "N": "NEW",
            "I": "INC",
            "S": "SRC",
            "Q": "EXT"
            }
        
        #some locations have a suite 
        suite = (sheet.cell(i, 6).value if sheet.cell(i, 6).value != None else None)
        if suite != None:
            addr = "{} {} {} {}".format(str(sheet.cell(i, 4).value), str(sheet.cell(i, 5).value),
                                        suite, str(sheet.cell(i, 7).value))
        else:
            addr = "{} {} {}".format(str(sheet.cell(i, 4).value), str(sheet.cell(i, 5).value), str(sheet.cell(i, 7).value))
        
        # change based on what you're looking at (0 or 1) 
        if sheet.cell(i, 30).value == 1 and sheet.cell(i, 31).value == None:
            # print bunch of stuff to console
            p11 = threading.Thread(target = naics_web_lookup, args = (naics_set, addr))
            p11.start()
            print("{} {}: {}".format(i, addr, da))
            print("{}: {}".format(sheet.cell(i, 2).value, str(sheet.cell(i, 3).value)))
            print(str(sheet.cell(i, 23).value))
            print("Building type: 1 {} {}%, 2 {} {}%, 3 {} {}%".format(sheet.cell(i, 10).value, sheet.cell(i, 13).value,
                                                                          sheet.cell(i, 11).value, sheet.cell(i, 14).value,
                                                                          sheet.cell(i, 12).value, sheet.cell(i, 15).value))
            print("{}\nBuildings: {}, sq.ft: {}; yr built: {}, yr moved in: {}, date surveyed: {}".format(
                sheet.cell(i, 16).value, sheet.cell(i, 18).value, sheet.cell(i, 17).value,
                sheet.cell(i, 19).value, sheet.cell(i, 21).value, sheet.cell(i, 22).value))
            p11.join()
            
            #loop here if you search terms, otherwise break out of loop
            exit_cond = False
            while not exit_cond:
                temp = str(input("CEC (1-6, N(ew), I(nc), S(earch), Q(uit): ")).upper()
                while temp not in dd.keys():  # input must be in the keys of dictionary dd
                    temp = str(input("Invalid, please enter 1-6, N(ew), I(nc), S(earch), Q(uit): ")).upper()
                val = dd[temp]  # return input as a command
                if val == "SRC": # search term(s), will not only search for entries with both words
                    curr = input("Src: ")
                    for key in naics_desc:
                        a = naics_desc[key]
                        b = curr.split()
                        for desc in a:
                            for word in b:
                                if word in desc.lower():
                                    print("{} {}".format(key, desc))
                    exit_cond = False
                
                elif val == "EXT": # break out to quit
                    break
                elif val == "VER": # at least one of the codes is correct, 
                                   # ver needs to be written specially to show which sources are correct
                    zed = [val]
                    sheet.cell(i, 32).value = da[temp]
                    for num in da.keys():
                        if da[num] == da[temp]:
                            zed.append(num)
                    temp3 = ''.join(zed)
                    sheet.cell(i, 31).value = temp3
                    exit_cond = True
                    
                elif val == "NEW": # determined that none of given codes are correct
                    sheet.cell(i, 31).value = val 
                    temp2 = checker(int(input("New CEC: ")), code_list)
                    while temp2 == 0:
                        temp2 = checker(int(input("Invalid CEC: ")), code_list)
                    sheet.cell(i, 32).value = temp2
                    sheet.cell(i, 34).value = input("Note: ") # must have a note
                    exit_cond = True
            
                else: # inconclusive site: empty buildings, no info online, etc
                    sheet.cell(i, 31).value = val
                    sheet.cell(i, 32).value = ""
                    sheet.cell(i, 34).value = input("Note: ") # must have a note
                    exit_cond = True
            print()

            action = input("Quit(Q), Notes(N), Save(S), Undo Later(U)? ").upper()
            if action == "Q": # close chrome windows, break to exit
                os.system('taskkill /im chrome.exe /f')
                break
            elif action == "N": # add a note to the site, if you wrote a note earlier it will be overwritten
                sheet.cell(i, 34).value = input("note... ")
            elif action == "S": # manual save, autosave counter resets
                book.save(editorFile)
                save_action = 1
            elif action == "U": # undo entry by clearing what you inputted (does not clear notes)
                sheet.cell(i, 31).value = None
                sheet.cell(i, 32).value = None
                
            print()
            save_counter += 1 # increment to autosave
            
        # built-in autosave, change <if save_counter == {}> to make it autosave after a specified amount 
        if save_counter == 10 and save_action == 0:
            ab = threading.Thread(target = book.save, args = (editorFile,))
            ba = threading.Thread(target = os.system, args = ('taskkill /im chrome.exe /f',))
            ab.start()
            ba.start()
            print()
            print("////////////////////")
            print("//// Autosaving ////")
            print("////////////////////")
            print()
            save_counter = 0
            ab.join()
            ba.join()
        elif save_action == 1:
            save_counter = 0
    book.save(editorFile)
    print("\nDone!")
        

if __name__ == "__main__":
    # use your own filenames, beginning and endpoints.
    ceushelper(2, 13000, 'C:\\Users\\Ksuzuki\\Desktop\\naics_13k 08-19-2019 ALL 4000 - Copy.xlsx', 'C:\\users\\Ksuzuki\\Desktop\\2017_NAICS_Structure.xlsx', 'C:\\users\\Ksuzuki\\Desktop\\2017_NAICS_Index_File.xlsx')
